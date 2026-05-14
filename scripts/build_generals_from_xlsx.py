"""xlsx 武将シートから武将エントリを自動生成し、generals.yaml に追記する。

- 試作9体（manual entries）は触らない。
- 残りの武将は xlsx 由来のデータ（兵種・配置・備考）から自動生成。
- 備考列の cap パターンを正規表現で抽出（不完全な場合は raw 備考 を effect に保持）。
- 玄武将は base_general_id_guess に推定覚醒元名を格納。
"""
import re
import sys
import yaml
from pathlib import Path
from openpyxl import load_workbook

# プロジェクトルート（scripts/ の親ディレクトリ）。
ROOT = Path(__file__).resolve().parent.parent
XLSX = str(ROOT / '3_20251224.xlsx')
YAML_PATH = str(ROOT / 'generals.yaml')

OVERLAP_NAMES = {'張遼', '于禁', '関羽', '馬超', '張寧姫', '賈逵', '玄趙雲'}
CATEGORY_TO_TROOP = {'歩兵':'歩兵', '馬兵':'騎兵', '弓兵':'弓兵', '戦車':'戦車', '共通':'万能'}

# scope 推定: 備考に集結・連撃・抵抗・全兵士・味方軍団等のキーワードがあれば rally、
# 軍団・部隊比率・所属軍団であれば legion、それ以外は scope:unknown 扱い→legion 既定
RALLY_KW = ['集結', '味方集結', '集結軍団', '共有', '非共有', '・プレイヤー数差', '味方全', '敵全']
ARRAY_KW = ['角陣', '副陣']
LEGION_KW = ['今いる軍団', '所属軍団', '所属武将軍団', '軍団内', '軍団の', '部隊', '本軍団']

CAP_PAT_BASIC = re.compile(r'（上限\s*(-?\d+(?:\.\d+)?)\s*[%％](\s*[・,、]\s*(共有|非共有|累加可))?\s*(?:[・,、]\s*1\s*スキル上限\s*(\d+)\s*[%％])?[）)]')
NON_STACK_PAT = re.compile(r'（累加不可）')
NO_CAP_PAT = re.compile(r'（上限\s*無し|上限なし）')
BOUND_PAT = re.compile(r'\(?下限\s*(\d+)\s*[%％]\s*[,、]\s*上限\s*(\d+)\s*[%％]\)?')
PCT_VALUE = re.compile(r'(-?\d+(?:\.\d+)?)\s*[%％]')


def split_skills(remarks: str):
    """備考を「、」「，」「。」で分割する。【装備】はサブ要素として残す。"""
    # まず【装備】の前で分割（装備効果は独立スキル扱い）
    chunks = []
    parts = re.split(r'(【装備】)', remarks)
    buf = ''
    for part in parts:
        if part == '【装備】':
            if buf.strip():
                chunks.append(('main', buf.strip().rstrip('、，。 ')))
            buf = ''
            chunks.append(('equipment_marker', ''))
        else:
            buf += part
    if buf.strip():
        # 最後の chunk が装備直後なら equipment、それ以外は main
        kind = 'equipment' if chunks and chunks[-1][0] == 'equipment_marker' else 'main'
        if chunks and chunks[-1][0] == 'equipment_marker':
            chunks.pop()
        chunks.append((kind, buf.strip().rstrip('、，。 ')))

    # 各 chunk を「、」「，」で更に分割（ただし括弧内のカンマでは分割しない）
    skills_raw = []
    for kind, text in chunks:
        for s in split_outside_parens(text):
            s = s.strip()
            if s:
                skills_raw.append((kind, s))
    return skills_raw


def split_outside_parens(text: str):
    """全角/半角括弧の外側でのみ「、」「，」で分割する。"""
    parts = []
    buf = []
    depth = 0
    for ch in text:
        if ch in '（(':
            depth += 1
            buf.append(ch)
        elif ch in '）)':
            depth = max(0, depth - 1)
            buf.append(ch)
        elif ch in '、，' and depth == 0:
            parts.append(''.join(buf))
            buf = []
        else:
            buf.append(ch)
    if buf:
        parts.append(''.join(buf))
    return parts


def guess_scope(text: str) -> str:
    if any(kw in text for kw in ARRAY_KW):
        return 'array'
    if any(kw in text for kw in RALLY_KW):
        return 'rally'
    if any(kw in text for kw in LEGION_KW):
        return 'legion'
    # 「（上限X%）」表記は典型的に集結系 → rally に倒す
    if '上限' in text and '%' in text:
        return 'rally'
    # 「累加不可」「反射」「ダメ X%増」などはパッシブ寄り = self
    return 'self'


def parse_one_skill(kind: str, text: str):
    """単一スキル文字列を skill dict に変換。"""
    scope = guess_scope(text)
    unit = 'percent' if '%' in text or '％' in text else 'value'

    skill = {
        'name': text[:30],
        'type': 'buff' if '増' in text else ('debuff' if '減' in text else 'passive'),
        'scope': scope,
        'unlocked_at': 0,
        'effect': text,
        'unit': unit,
        'value': {
            'by_breakthrough': [None, None, None, None, None],
            'confidence': 'low',
        },
    }

    # 装備限定なら unlocked_at_equipment 設定（構造化フィールドが正準）。
    # 既定値は 4（装備4で解放）。name にはプレフィックスを付けない。
    if kind == 'equipment':
        skill['unlocked_at_equipment'] = 4

    # 主効果数値の抽出
    nums = PCT_VALUE.findall(text)
    if nums:
        # 最初の数値を採用、増/減で符号
        v = float(nums[0])
        if '減' in text and v > 0:
            v = -v
        # int化
        skill['value']['by_breakthrough'] = [None, None, None, None, v]
        skill['value']['confidence'] = 'medium'  # xlsx由来

    # 累加不可
    if NON_STACK_PAT.search(text):
        skill['stacking'] = 'non_stackable'

    # 上限なし
    if NO_CAP_PAT.search(text):
        skill['cap'] = {'rule': 'no_cap', 'note': 'xlsx備考で上限無しと明示'}
        return skill

    # 上下限ペア
    m = BOUND_PAT.search(text)
    if m:
        floor_v = int(m.group(1))
        cap_v = int(m.group(2))
        skill['cap'] = {
            'rule': 'no_cap',  # bounded系の汎用rule。後で適切なrule追加するなら差し替え
            'cap_value': cap_v,
            'floor_value': floor_v,
            'note': 'xlsx備考の下限/上限ペア',
        }
        return skill

    # 標準上限パターン
    m = CAP_PAT_BASIC.search(text)
    if m:
        cap_v = float(m.group(1))
        modifier = m.group(3)  # 共有/非共有/累加可
        per_skill = m.group(4)
        if '減' in text and cap_v > 0:
            cap_v = -cap_v
        skill['cap'] = {
            'rule': 'no_cap',   # ルール名は汎用。具体ルール決定は手作業（cap_rules辞書を別途参照する場合は後でマージ）
            'cap_value': cap_v,
            'note': 'xlsx備考由来の固定%上限',
        }
        if modifier == '共有':
            skill['cap']['shared'] = True
        elif modifier == '非共有':
            skill['cap']['shared'] = False
        if per_skill:
            skill['cap']['per_skill_cap'] = int(per_skill)

    return skill


def parse_remarks(remarks: str):
    """備考全体を skills の配列に変換。"""
    if not remarks:
        return []
    parts = split_skills(remarks)
    return [parse_one_skill(kind, text) for kind, text in parts if text]


def main():
    with open(YAML_PATH, 'r', encoding='utf-8') as f:
        doc = yaml.safe_load(f)
    # 冪等性: 過去にこのスクリプトで生成したエントリ（xlsx_remarks 付き or sources が xlsx 開始）を除去
    before = len(doc['generals'])
    doc['generals'] = [
        g for g in doc['generals']
        if 'xlsx_remarks' not in g and not (g.get('sources') and len(g['sources']) == 1 and g['sources'][0].startswith('xlsx:'))
    ]
    purged = before - len(doc['generals'])
    if purged:
        print(f'[purge] 旧自動生成 {purged}件を除去', file=sys.stderr)
    existing_ids = {g['id'] for g in doc['generals']}
    existing_names = {g['name'] for g in doc['generals']}
    print(f'[既存] 手動分 {len(doc["generals"])}件', file=sys.stderr)

    wb = load_workbook(XLSX, data_only=True)
    ws = wb['武将']
    current_cat = None
    new_generals = []
    skip_names = {'小計', '合計', '武将'}

    for r_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
        a = row[0] if len(row) > 0 else None
        b = row[1] if len(row) > 1 else None
        c = row[2] if len(row) > 2 else None
        w = row[22] if len(row) > 22 else None
        if isinstance(a, str) and a in CATEGORY_TO_TROOP:
            current_cat = a
            continue
        if not b or not isinstance(b, str):
            continue
        name = b.strip()
        if name in skip_names or not name:
            continue
        if name in OVERLAP_NAMES or name in existing_names:
            continue
        if not current_cat:
            continue

        rarity = '玄' if name.startswith('玄') else '橙'
        troop = CATEGORY_TO_TROOP[current_cat]

        slot_stat = None
        if isinstance(c, str):
            c_str = c.strip()
            if c_str in ('ー', '', None):
                slot_stat = None
            elif '/' in c_str:
                slot_stat = [('知性' if s.strip() == '知力' else s.strip()) for s in c_str.split('/')]
            elif c_str in ('統率', '武力', '知性', '知力'):
                slot_stat = '知性' if c_str == '知力' else c_str

        gid = name
        if gid in existing_ids:
            gid = f'{name}_xlsx'
        existing_ids.add(gid)

        remarks = (w or '').strip() if isinstance(w, (str, type(None))) else str(w).strip()
        skills = parse_remarks(remarks) if remarks else []

        entry = {
            'id': gid,
            'name': name,
            'yomi': None,
            'base_rarity': rarity,
            'troop_type': troop,
            'slot_stat': slot_stat,
            'role_tags': [],
            'skills': skills,
            'seal_bonus': {
                'by_seal': [None] * 6,
                'effect_note': '未確認',
                'confidence': 'unknown',
            },
            'sources': [f'xlsx:3_20251224.xlsx#武将!B{r_idx}'],
        }
        if rarity == '玄':
            entry['base_general_id_guess'] = name[1:]
        if remarks:
            entry['xlsx_remarks'] = remarks
        new_generals.append(entry)

    print(f'[抽出] 新規 {len(new_generals)}件', file=sys.stderr)

    # 玄武将の base_general_id 解決（同 YAML 内に覚醒元名があれば紐付け）
    all_entries = doc['generals'] + new_generals
    name_to_id = {g['name']: g['id'] for g in all_entries}
    for g in new_generals:
        guess = g.pop('base_general_id_guess', None)
        if guess and guess in name_to_id:
            g['base_general_id'] = name_to_id[guess]

    doc['generals'].extend(new_generals)
    doc['meta']['data_version'] = '0.4-full'
    doc['meta']['last_updated'] = '2026-05-14'

    with open(YAML_PATH, 'w', encoding='utf-8') as f:
        yaml.safe_dump(doc, f, allow_unicode=True, sort_keys=False, width=200)
    print(f'[完了] 合計 {len(doc["generals"])}件 → {YAML_PATH}', file=sys.stderr)


if __name__ == '__main__':
    main()
